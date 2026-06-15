
from __future__ import annotations

import json
import re
import hashlib
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

from openpyxl import load_workbook

MESES = {
    "enero": 1, "febrero": 2, "marzo": 3, "abril": 4, "mayo": 5, "junio": 6,
    "julio": 7, "agosto": 8, "septiembre": 9, "setiembre": 9, "octubre": 10,
    "noviembre": 11, "diciembre": 12,
}

CANONICAL_FIELDS = [
    "source_file", "sheet_name", "nivel", "unidad", "area", "mes", "mes_numero",
    "semana_mes", "semana_lectiva", "semana_calendario", "estado_calendario",
    "eje", "contenido", "foco", "contexto", "momento1", "momento2", "momento3",
    "momento4", "observaciones", "original_json", "row_hash",
]


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    return str(value).strip()


def normalize_header(value: Any) -> str:
    text = clean_text(value).lower()
    text = re.sub(r"\s+", " ", text)
    text = text.replace("á", "a").replace("é", "e").replace("í", "i").replace("ó", "o").replace("ú", "u")
    text = text.replace("/", " ").replace("-", " ").replace("_", " ")
    text = re.sub(r"[^a-z0-9ñ ]+", "", text)
    return re.sub(r"\s+", " ", text).strip()


def detect_source_metadata(filename: str) -> Dict[str, str]:
    f = filename.lower()
    if "inicial" in f:
        nivel = "Nivel Inicial"
    elif "primaria" in f:
        nivel = "Nivel Primario"
    elif "secundaria" in f or "ciclo_basico" in f or "ciclo basico" in f:
        nivel = "Nivel Secundario - Ciclo Básico"
    else:
        nivel = ""

    if "matematica" in f and "geometria" in f:
        area = "Matemática y Geometría"
    elif "matematica" in f:
        area = "Matemática"
    elif "lengua" in f or "lenguaje" in f or "literatura" in f:
        area = "Lengua / Prácticas del Lenguaje"
    else:
        area = ""
    return {"nivel": nivel, "area": area}


def find_header_row(ws, max_scan_rows: int = 12) -> Optional[int]:
    """Busca una fila de encabezado flexible en las primeras filas de una hoja."""
    max_row = ws.max_row or 0
    max_col = ws.max_column or 0
    for r in range(1, min(max_row, max_scan_rows) + 1):
        values = [clean_text(ws.cell(r, c).value) for c in range(1, min(max_col, 25) + 1)]
        joined = " | ".join(values).lower()
        has_week = "semana" in joined
        has_content = "contenido" in joined or "recorte" in joined or "eje" in joined
        has_unit = "grado" in joined or "sala" in joined or "año" in joined or "nivel" in joined
        if has_week and has_content and has_unit:
            return r
    return None


def selected_sheets(workbook) -> List[str]:
    """Prioriza la hoja general para evitar duplicar registros por grado/sala/año."""
    names = workbook.sheetnames
    general = [n for n in names if "cronograma" in n.lower() and "general" in n.lower()]
    if general:
        return general
    return names


def row_to_dict(headers: List[str], row: Tuple[Any, ...]) -> Dict[str, Any]:
    result = {}
    for h, v in zip(headers, row):
        if h:
            result[clean_text(h)] = v
    return result


def pick(raw: Dict[str, Any], *candidates: str) -> str:
    normalized = {normalize_header(k): v for k, v in raw.items()}
    for cand in candidates:
        key = normalize_header(cand)
        if key in normalized:
            return clean_text(normalized[key])
    # búsqueda por inclusión para columnas similares
    for cand in candidates:
        c = normalize_header(cand)
        for k, v in normalized.items():
            if c in k or k in c:
                return clean_text(v)
    return ""


def month_number(mes: str) -> Optional[int]:
    return MESES.get(clean_text(mes).lower())


def infer_unidad(raw: Dict[str, Any], sheet_name: str) -> str:
    unidad = pick(raw, "Grado", "Sala", "Año")
    if unidad:
        return unidad
    s = sheet_name.strip()
    if re.search(r"sala", s, flags=re.I):
        return s.replace("_", " ")
    if re.search(r"grado", s, flags=re.I):
        return s.replace("_", " ")
    if re.search(r"año|ano", s, flags=re.I):
        return s.replace("_", " ")
    return ""


def normalize_record(raw: Dict[str, Any], filename: str, sheet_name: str) -> Dict[str, Any]:
    meta = detect_source_metadata(filename)
    nivel = pick(raw, "Nivel") or meta["nivel"]
    area = pick(raw, "Espacio curricular", "Área/Campo", "Area Campo", "Área") or meta["area"]
    unidad = infer_unidad(raw, sheet_name)

    mes = pick(raw, "Mes")
    semana_mes = pick(raw, "Semana del mes")
    semana_lectiva = pick(raw, "Semana lectiva")
    semana_calendario = pick(raw, "Semana calendario")
    estado = pick(raw, "Estado", "Estado calendario")
    eje = pick(raw, "Eje temático", "Eje temático del Diseño Curricular")
    contenido = pick(raw, "Contenido / recorte curricular semanal", "Contenido / recorte semanal", "Contenido / recorte curricular", "Contenido")
    foco = pick(raw, "Foco de alfabetización (Plan CUMBRE)", "Foco de alfabetización")
    contexto = pick(raw, "Contexto o situación de enseñanza")
    momento1 = pick(raw, "Momento 1 - Personaliza / Activación", "Momento 1 - PERSONALIZA / Activación", "Momento 1")
    momento2 = pick(raw, "Momento 2 - Aprende / Recurso central", "Momento 2 - APRENDE / Recurso central", "Momento 2")
    momento3 = pick(raw, "Momento 3 - Crea / Práctica guiada", "Momento 3 - CONVERSA-CREA / Práctica guiada", "Momento 3")
    momento4 = pick(raw, "Momento 4 - Comparte / Producción", "Momento 4 - COMPARTE / Producción/Evidencia", "Momento 4")
    observaciones = pick(raw, "Observaciones para Moodle", "Observación curricular", "Observaciones Moodle", "Criterio de seguimiento docente")

    normalized = {
        "source_file": filename,
        "sheet_name": sheet_name,
        "nivel": nivel,
        "unidad": unidad,
        "area": area,
        "mes": mes,
        "mes_numero": month_number(mes),
        "semana_mes": semana_mes,
        "semana_lectiva": semana_lectiva,
        "semana_calendario": semana_calendario,
        "estado_calendario": estado,
        "eje": eje,
        "contenido": contenido,
        "foco": foco,
        "contexto": contexto,
        "momento1": momento1,
        "momento2": momento2,
        "momento3": momento3,
        "momento4": momento4,
        "observaciones": observaciones,
    }
    original_json = json.dumps({k: clean_text(v) for k, v in raw.items()}, ensure_ascii=False)
    fingerprint_base = "|".join(clean_text(normalized.get(k)) for k in [
        "nivel", "unidad", "area", "mes", "semana_lectiva", "semana_calendario", "eje", "contenido"
    ])
    normalized["original_json"] = original_json
    normalized["row_hash"] = hashlib.sha256(fingerprint_base.encode("utf-8")).hexdigest()
    return normalized


def load_curriculum_from_excel(path: str | Path) -> List[Dict[str, Any]]:
    path = Path(path)
    wb = load_workbook(path, read_only=False, data_only=True, keep_vba=path.suffix.lower() == ".xlsm")
    records: List[Dict[str, Any]] = []
    for sheet_name in selected_sheets(wb):
        ws = wb[sheet_name]
        header_row = find_header_row(ws)
        if not header_row:
            continue
        max_col = ws.max_column or 0
        headers = [clean_text(ws.cell(header_row, c).value) for c in range(1, max_col + 1)]
        for row in ws.iter_rows(min_row=header_row + 1, max_col=max_col, values_only=True):
            raw = row_to_dict(headers, row)
            # Ignorar filas vacías o de subtítulos.
            if not any(clean_text(v) for v in raw.values()):
                continue
            rec = normalize_record(raw, path.name, sheet_name)
            if rec["contenido"] or rec["eje"]:
                records.append(rec)
    return records


def analyze_workbook(path: str | Path) -> List[Dict[str, Any]]:
    path = Path(path)
    wb = load_workbook(path, read_only=False, data_only=True, keep_vba=path.suffix.lower() == ".xlsm")
    info = []
    for ws in wb.worksheets:
        header_row = find_header_row(ws)
        headers = []
        data_rows = 0
        if header_row:
            max_col = ws.max_column or 0
            headers = [clean_text(ws.cell(header_row, c).value) for c in range(1, max_col + 1) if clean_text(ws.cell(header_row, c).value)]
            for row in ws.iter_rows(min_row=header_row + 1, max_col=max_col, values_only=True):
                if any(clean_text(v) for v in row):
                    data_rows += 1
        info.append({
            "archivo": path.name,
            "hoja": ws.title,
            "fila_encabezado": header_row or "No detectada",
            "columnas_detectadas": ", ".join(headers),
            "filas_datos": data_rows,
        })
    return info
